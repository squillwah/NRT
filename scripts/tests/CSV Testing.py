import csv
import openpyxl
from openpyxl import Workbook

def write_csv(inputted_dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output from Models"

    for i in range(len(inputted_dict)):
        new_sheet = wb.create_sheet(title="Test" + str(i))
        headers = ["Overall", "Author", "Journal", "Publish Date", "Author Order", "Publisher", "Percentage of Confidence", "Model"]

        new_sheet.append(headers)
        row_values = [inputted_dict[key] for key in headers]
        new_sheet.append(row_values)

    wb.save(filename="output.xlsx")

# add title of the citation

def write_xlsx(inputted_data):

    wb = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl so it doesn't leave an empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 2. Iterate through the list of dictionaries
    for index, inputted_dict in enumerate(inputted_data):
        # Dynamically name each sheet (e.g., Sheet_1, Sheet_2)
        sheet_name = f"Model_{index + 1}"
        ws = wb.create_sheet(title=sheet_name)

        # 3. Extract headers and values for THIS dictionary specifically
        headers = list(inputted_dict.keys())
        values = [inputted_dict[key] for key in headers]

        # 4. Append the headers as row 1, and values as row 2
        ws.append(headers)
        ws.append(values)

    wb.save(filename="output.xlsx")


data = [
    {"Overall": "Fake", "Author": "Real", "Journal": "Real", "Publish Date": "Fake", "Author Order": "Real", "Publisher": "Real", "Percentage of Confidence": 95.00001, "Model": "Chat-GPT"},
    {"Overall": "Fake", "Author": "Real", "Journal": "Real", "Publish Date": "Fake", "Author Order": "Real", "Publisher": "Real", "Percentage of Confidence": 95.00001, "Model": "Gemini"}
]

if __name__ == "__main__":
    write_xlsx(data)
#
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Test 1"
#
# for i in range(len(data)):
#     new_sheet = wb.create_sheet(title="Test" + str(i))
#     headers = ["Overall", "Author", "Journal", "Publish Date", "Author Order", "Publisher", "Percentage of Confidence", "Model"]
#
#     new_sheet.append(headers)
#     for item in data:
#         row_values = [item[key] for key in headers]
#         new_sheet.append(row_values)
#
#
# wb.save(filename= "output.xlsx")

# make a final column of what the correct answer should be
# need to create a full excel

