import csv
import openpyxl
import json
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re


# def write_csv(inputted_json):
#
#     with open(inputted_json, 'r', encoding='utf-8') as file:
#         data = json.load(file)
#
#     json_results = data["results"]
#     json_keynames = data["components"]
#     journal_id = str(data["id"])
#
#     json_keynames.insert(0, "model")
#     json_keynames.remove("REFERENCE")
#     json_keynames.remove("url_abstract")
#     json_keynames.remove("url_direct")
#     json_keynames.insert(1, "REFERENCE")
#
#
#     with open(journal_id + ".csv", "w") as csv_f:
#         writer = csv.writer(csv_f)
#
#         writer.writerow(json_keynames)
#
#         for model in json_results:
#             model_results = json_results[model]
#             for form in model_results:
#                 if model_results[form] is not None:
#                     component_data = model_results[form]
#                     csv_f.write(model + ",")
#                 else:
#                     continue
#                 for component in component_data:
#                     csv_f.write(str(component_data[component]["validity"]))
#                     csv_f.write(",")
#                 csv_f.write("\n")


def write_all_csv(inputted_json):
    with open(inputted_json, 'r', encoding='utf-8') as file:
        data = json.load(file)

        for paper in data.values():

            json_results = paper["results"]
            json_keynames = paper["components"]
            journal_id = str(paper["id"])

            json_keynames.insert(0, "model")
            json_keynames.remove("REFERENCE")
            json_keynames.remove("url_abstract")
            json_keynames.remove("url_direct")
            json_keynames.insert(1, "REFERENCE")

            with open(journal_id + ".csv", "w") as csv_f:
                writer = csv.writer(csv_f)

                writer.writerow(json_keynames)

                for model in json_results:
                    model_results = json_results[model]
                    for form in model_results:
                        if model_results[form] is not None:
                            component_data = model_results[form]
                            csv_f.write(model + ",")
                        else:
                            continue
                        for component in component_data:
                            csv_f.write(str(component_data[component]["validity"]))
                            csv_f.write(",")
                        csv_f.write("\n")

# def make_average_csv(inputted_csv):
#     with open(inputted_csv, 'r', encoding='utf-8') as file:
#         reader = csv.reader(file)
#         csv_lines = [row for row in reader]





def write_xlsx():

    wb = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl so it doesn't leave an empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    inputted_csv = input("Please enter the csv file path: ")

    avg_total_divisor_acc = 0
    avg_total_divisor_conf = 0

    avg_accuracy = []
    avg_confidence = []
    avg_accuracy_calc = []
    avg_confidence_calc = []
    models = []

    while inputted_csv != "end":


        burner_list = []

        with open(inputted_csv, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            csv_id = re.sub(r'[^a-zA-Z0-9]', '', inputted_csv)
            csv_id = csv_id.replace("CUsersmastePycharmProjectsNRTsrc", '')
            csv_id = csv_id.replace("csv", '')
            csv_lines = [row for row in reader]




            if "accuracy" in csv_id:

                avg_total_divisor_acc += 1

                if avg_total_divisor_acc <= 1:
                    avg_accuracy.append(csv_lines[0])
                    for k in range(1, len(csv_lines)):
                        avg_accuracy_calc.append(csv_lines[k][2:])

                    for l in range(1, len(csv_lines)):
                        models.append(csv_lines[l][0])

                for k in range(1, len(csv_lines)):
                    burner_list.append(csv_lines[k][2:])

                for i, row in enumerate(burner_list):
                    for j , cell in enumerate(row):
                        if "True" in cell:
                            if avg_total_divisor_acc > 1:
                                avg_accuracy_calc[i][j] += 1
                            else:
                                avg_accuracy_calc[i][j] = 1
                        elif "False" in cell:
                            if avg_total_divisor_acc > 1:
                                avg_accuracy_calc[i][j] += 0
                            else:
                                avg_accuracy_calc[i][j] = 0
                        else:
                            continue

            elif "confidence" in csv_id:

                avg_total_divisor_conf += 1

                if avg_total_divisor_conf <= 1:
                    avg_confidence.append(csv_lines[0])

                    for k in range(1, len(csv_lines)):
                        avg_confidence_calc.append(csv_lines[k][2:])

                    for l in range(1, len(csv_lines)):
                        models.append(csv_lines[l][0])

                for k in range(1, len(csv_lines)):
                    burner_list.append(csv_lines[k][2:])

                for i, row in enumerate(burner_list):
                    for j, cell in enumerate(row):
                        if avg_total_divisor_conf > 1:
                            avg_confidence_calc[i][j] += float(cell)
                        else:
                            avg_confidence_calc[i][j] = float(cell)



            xlsx_header = csv_lines[0]

            ws = wb.create_sheet(title=csv_id)

            ws.append(xlsx_header)

            header_font = Font(name="Calibri", size=11, bold=True, color="000000")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
            header_align = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for row in csv_lines[1:]:
                ws.append(row)


            for col in ws.columns:
                max_len = 0
                column = col[0].column
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))

                # Set width with buffer, min width 10
                adjusted_width = (max_len + 3)
                ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        inputted_csv = input("Enter another csv file path or enter 'end': ")


    xlsx_header.remove(xlsx_header[1])

    avg_accuracy_list = create_avg_list(avg_accuracy_calc, xlsx_header, avg_total_divisor_acc, models)
    avg_confidence_list = create_avg_list(avg_confidence_calc, xlsx_header, avg_total_divisor_conf, models)

    # ---------------- Make avg_accuracy sheet --------------------------------

    sheet_name = "avg_accuracy"

    avg_header = avg_accuracy_list[0]

    ws = wb.create_sheet(title=sheet_name)

    ws.append(avg_header)

    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in avg_accuracy_list[1:]:
        ws.append(row)

    for col in ws.columns:
        max_len = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        # Set width with buffer, min width 10
        adjusted_width = (max_len + 3)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # ---------------- Make avg_confidence sheet --------------------------------

    sheet_name = "avg_confidence"

    avg_header = avg_confidence_list[0]

    ws = wb.create_sheet(title=sheet_name)

    ws.append(avg_header)

    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in avg_confidence_list[1:]:
        ws.append(row)

    for col in ws.columns:
        max_len = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        # Set width with buffer, min width 10
        adjusted_width = (max_len + 3)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


    wb.save(filename="full_results.xlsx")


def create_avg_list(value_list, headers_list, divisor_factor, models):
    for i, element in enumerate(value_list):
        for k, cell in enumerate(element):
            value_list[i][k] = cell / divisor_factor
        value_list[i].insert(0, models[i])


    value_list.insert(0, headers_list)

    return value_list
# def make_sheet(input_list, sheet_name):
#     avg_header = input_list[0]
#
#     ws = wb.create_sheet(title=sheet_name)
#
#     ws.append(avg_header)
#
#     header_font = Font(name="Calibri", size=11, bold=True, color="000000")
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
#     header_align = Alignment(horizontal="center", vertical="center")
#
#     for cell in ws[1]:
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = header_align
#
#     for row in input_list[1:]:
#         ws.append(row)
#
#     for col in ws.columns:
#         max_len = 0
#         column = col[0].column
#         for cell in col:
#             if cell.value:
#                 max_len = max(max_len, len(str(cell.value)))
#
#         # Set width with buffer, min width 10
#         adjusted_width = (max_len + 3)
#         ws.column_dimensions[get_column_letter(column)].width = adjusted_width


    # # 2. Iterate through the list of dictionaries
    # for i, dict_list in enumerate(inputted_data):
    #
    #     ws = wb.create_sheet(title=pmid_list[i])
    #
    #     headers = list(dict_list[0].keys())
    #     ws.append(headers)
    #     for inputted_dict in dict_list:
    #
    #         # Extract headers and values for THIS dictionary specifically
    #
    #         values = [inputted_dict[key] for key in headers]
    #
    #         # Append the headers as row 1, and values as row 2
    #
    #         ws.append(values)
    #
    # wb.save(filename="output.xlsx")


test_data = [[1, 1, 1, 1, 0, 0, 0, 1, 1, 0, 0, 0], [0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0]]
test_avg_accuracy = ['model', 'format', 'authors', 'title', 'journal_name', 'journal_volume', 'journal_issue', 'journal_page', 'elocator', 'publication_date', 'doi', 'pmcid', 'pmid', 'REFERENCE']
test_models = ['deepseek/deepseek-v4-flash', 'nvidia/nemotron-3-super-120b-a12b:free', 'xiaomi/mimo-v2.5', 'google/gemini-3-flash-preview']



pmids = ["65sd4f3654f", "4d441dsf44"]

if __name__ == "__main__":
    # create_avg_list(test_data, test_avg_accuracy, 1, test_models)
    write_xlsx()
    # write_all_csv("all.json")
